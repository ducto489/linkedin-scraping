import time
import os
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from gologin import GoLogin
from sys import platform
import openpyxl


def init_driver():
    options = Options()
    # options.add_argument('--headless')
    # options.add_argument('--no-sandbox')
    options.add_argument("--disable-dev-shm-usage")

    # init driver until success
    while True:
        try:
            driver = webdriver.Chrome(
                service=Service(ChromeDriverManager().install()), options=options
            )
            break
        except:
            pass

    return driver


def get_google_search_results(driver, query):
    driver.get(
        "https://www.google.com/search?q=site%3Alinkedin.com%2Fin%2F+{}".format(
            "+".join(query.split())
        )
    )
    return driver


def saving_to_csv(df, file_name):
    df.to_csv(file_name, index=False)


def init_worksheet(file_name):
    try:
        workbook = openpyxl.load_workbook(file_name)
    except:
        workbook = openpyxl.Workbook()
        workbook.save(file_name)

    return workbook


def get_company_list(file_name, sheet_name):
    try:
        workbook = openpyxl.load_workbook(file_name)
    except:
        raise Exception(f"File {file_name} not found")

    worksheet = workbook[sheet_name]

    # make sure worksheet has at least 2 columns
    if worksheet.max_column < 2:
        raise Exception("Sheet file requires 2 columns for Company, Country and Status")

    company_list = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        # if col 3 is blank, then we will add the company to the list
        if row[2] is None and row[0] is not None:
            company_list[row[0].lower()] = company_list.get(row[0].lower(), [])
            company_list[row[0].lower()].append(row[1])

    return company_list


def save_hr_list(file_name, sheet_name, hr_list):
    
    try:
        workbook = openpyxl.load_workbook(file_name)
    except:
        # create file if not found
        workbook = openpyxl.Workbook()
        # workbook.save(file_name)

    worksheet = None

    try:
        worksheet = workbook[sheet_name]
    except:
        # print("Create new sheet")
        worksheet = workbook.create_sheet(sheet_name)

    last_row = worksheet.max_row
    for row, item in enumerate(hr_list, start=1):
        for col, value in enumerate(item.values(), start=1):
            worksheet.cell(row=last_row + row, column=col, value=value)
    # save the workbook
    workbook.save(file_name)


def mark_as_scraped(file_name, sheet_name, country, company):
    try:
        workbook = openpyxl.load_workbook(file_name)
    except:
        raise Exception(f"File {file_name} not found")

    worksheet = workbook[sheet_name]
    i = 2
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        # print(row[0], row[1])
        if row[0] is not None and row[0].lower() == country.lower() and row[1].lower() == company.lower():
            worksheet.cell(row=i, column=3, value="done")
            break
        i += 1

    workbook.save(file_name)


class LinkedinCrawler:
    def __init__(self, driver):
        self.driver = driver

    # Scroll to the bottom of the page to load more results
    def scroll_to_bottom(self, button_text):
        last_height = 0
        a = None
        while True:
            for i in driver.find_elements(By.TAG_NAME, "a"):
                if i.get_attribute("aria-label") == button_text:
                    a = i

            try:
                a.click()
                return False
            except:
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(1)

                new_height = driver.execute_script("return document.body.scrollHeight")

                if last_height == new_height:
                    return True
                else:
                    last_height = new_height

        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

    def parse_title(self, title, descriptions, jobtitle_keywords):
        parts = title.split(" - ")
        name, jobtitle, company, location = None, None, None, None

        if len(parts) >= 2:
            name = parts[0]
            if len(parts) == 2:
                company = parts[1]

                if any(keyword in company.lower() for keyword in jobtitle_keywords):
                    jobtitle = company
                    company = None
            else:
                jobtitle = parts[1] + ((" - " + parts[2]) if len(parts) == 4 else "")
                company = parts[2 if len(parts) == 3 else 3]
        else:
            raise Exception(f"Title not in expected format: {title}")

        # get the location based on the first line of the descriptions
        if len(descriptions) >= 2:
            descriptions = descriptions[:2]
            first_line = descriptions[0].text.split(" · ")
            location = first_line[0]

        # if we cant find jobtitle in the title, we will try to find it in the second part of descriptions, also
        # filter out if jobtitle is not in the `keywords`
        if (
            jobtitle is None
            or any(keyword in jobtitle.lower() for keyword in jobtitle_keywords)
            == False
        ):
            description = descriptions[-1].text

            for keyword in jobtitle_keywords:
                if keyword in description.lower():
                    jobtitle = keyword
                    break
                
        assert jobtitle is not None, f"Got error with name: {name}"

        return name, jobtitle, company, location

    def parse_linkedin_blocks(self, blocks, jobtitle_keywords, num_have_read=-1):
        data = []
        i = 0
        for block in blocks:
            if i <= num_have_read:
                i += 1
                continue

            children = block.find_elements(By.XPATH, "./*")[:2]
            href = children[0].find_element(By.TAG_NAME, "a").get_attribute("href")
            title = children[0].find_element(By.TAG_NAME, "h3").text
            description = children[1].find_elements(By.XPATH, "./*")

            i += 1

            try:
                name, jobtitle, company, location = self.parse_title(
                    title, description, jobtitle_keywords
                )

                if jobtitle is None:
                    continue
                
                print(jobtitle)

                data.append(
                    {
                        "Name": name,
                        "Job Title": jobtitle,
                        "Company": company,
                        "Location": location,
                        "Link": href,
                    }
                )
            except Exception as e:
                pass

        return data, i

    def get_df(
        self,
        query,
        jobtitle_keywords,
        button_text,
        class_name="N54PNb BToiNc cvP2Ce",
        loops=5,
    ):
        """
        Args:
            query (str): query to search for
            jobtitle_keywords (list of str): keywords to search for in job title
            button_text (_type_): text of button to click to load more results (in VN, it is 'Kết quả khác')
            class_name (str): class name of the block to get data
            loops (int): number of times to loop through the page to get data (because the later results are often not relevant)
        """
        self.driver = get_google_search_results(self.driver, query)
        time.sleep(1)
        data = []

        num_have_read = -1

        for epoch in range(loops):
            # print(f"Loop {epoch + 1}")
            if self.scroll_to_bottom(button_text):
                print("Reached end of page")
                break

            blocks = self.driver.find_elements(
                By.XPATH, f"//div[@class='{class_name}']"
            )
            parsed_data, num_have_read = self.parse_linkedin_blocks(
                blocks, jobtitle_keywords, num_have_read
            )

            # print(len(parsed_data))

            data.extend(parsed_data)

        return data


jobtitle_keywords = ["recruit", "hr", "human resources", "talent acquisition"]
excel_file = "linkedin_scrape.xlsx"
company_list = get_company_list(excel_file, "Company")

# cnt = 0
os.system("warp-cli disconnect")

print("Start scraping")

for country in company_list:
    for company in company_list[country]:
        # if cnt % 2 == 0:
        os.system("warp-cli connect")
        driver = init_driver()

        print(f"Scraping {company} in {country}")
        # with open(f"./log/{country}_{company}_log.txt", "w") as log_file:
        lc = LinkedinCrawler(driver)
        hr_list = lc.get_df(
            f"recruiter in company {company} location {country}",
            jobtitle_keywords,
            "Kết quả khác",
            loops=3,
        )

        save_hr_list(excel_file, country, hr_list)

        if len(hr_list) > 0:
            mark_as_scraped(excel_file, "Company", country, company)

        driver.close()
        os.system("warp-cli disconnect")
        

print("Scraping finished")

# driver.close()
driver.quit()